import customtkinter as ctk
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import tkinter as tk

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

facturas = []
contador_factura = 1

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Sistema de Facturación — Acueducto")
        self.resizable(True, True)
        self.minsize(700, 500)
        self.after(0, lambda: self.state("zoomed"))

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_header()
        self._build_form()
        self._build_table()
        self._build_footer()

    def _build_header(self):
        frame = ctk.CTkFrame(self, fg_color="#1F4E79", corner_radius=0, height=64)
        frame.grid(row=0, column=0, sticky="ew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_propagate(False)
        ctk.CTkLabel(frame, text="💧 EMPRESA DE ACUEDUCTO",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color="white").pack(side="left", padx=20, pady=14)
        ctk.CTkLabel(frame, text="Sistema de Facturación",
                     font=ctk.CTkFont(size=12),
                     text_color="#BDD7EE").pack(side="left")

    def _build_form(self):
        outer = ctk.CTkFrame(self, fg_color="transparent")
        outer.grid(row=1, column=0, sticky="ew", padx=20, pady=(14, 0))
        outer.grid_columnconfigure((1, 3, 5), weight=1)

        ctk.CTkLabel(outer, text="Nueva Factura",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color="#1F4E79").grid(row=0, column=0, columnspan=8,
                                                sticky="w", pady=(0, 8))

        labels = ["Cliente", "Largo (m)", "Ancho (m)"]
        self.entries = {}
        keys = ["cliente", "largo", "ancho"]

        for i, (lbl, key) in enumerate(zip(labels, keys)):
            ctk.CTkLabel(outer, text=lbl,
                         font=ctk.CTkFont(size=12)).grid(row=1, column=i*2,
                                                          sticky="w", padx=(0, 4))
            entry = ctk.CTkEntry(outer, placeholder_text=lbl)
            entry.grid(row=1, column=i*2+1, sticky="ew", padx=(0, 16))
            self.entries[key] = entry

        ctk.CTkButton(outer, text="＋ Registrar Factura", width=160,
                      fg_color="#1F4E79", hover_color="#2E75B6",
                      command=self._registrar).grid(row=1, column=6, padx=(8, 0))

                
        ctk.CTkButton(outer, text="🗑 Borrar Factura", width=160,
                      fg_color="#C0392B", hover_color="#922B21",
                      command=self._borrar).grid(row=1, column=7, padx=(8, 0))

    def _build_table(self):
        frame = ctk.CTkFrame(self, fg_color="transparent")
        frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=12)
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(frame, text="Facturas Registradas",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color="#1F4E79").grid(row=0, column=0, sticky="w", pady=(0, 6))

        container = ctk.CTkFrame(frame, fg_color="#F0F4F8", corner_radius=8)
        container.grid(row=1, column=0, sticky="nsew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(container, bg="#F0F4F8", highlightthickness=0)
        scrollbar_y = ctk.CTkScrollbar(container, command=self.canvas.yview)
        scrollbar_x = ctk.CTkScrollbar(container, orientation="horizontal",
                                        command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=scrollbar_y.set,
                               xscrollcommand=scrollbar_x.set)
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        self.canvas.grid(row=0, column=0, sticky="nsew")

        self.tabla_frame = tk.Frame(self.canvas, bg="#F0F4F8")
        self.canvas_window = self.canvas.create_window(
            (0, 0), window=self.tabla_frame, anchor="nw")

        self.tabla_frame.bind("<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", self._on_canvas_resize)

        self._render_tabla()

    def _on_canvas_resize(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)
        self._render_tabla()

    def _render_tabla(self):
        for widget in self.tabla_frame.winfo_children():
            widget.destroy()

        cols  = ["N° Factura", "Cliente", "Largo (m)", "Ancho (m)",
                 "Área (m²)", "Tarifa ($/m²)", "Total ($)"]
        pesos = [1, 3, 1, 1, 1, 2, 1]

        for j, peso in enumerate(pesos):
            self.tabla_frame.grid_columnconfigure(j, weight=peso)

        for j, col in enumerate(cols):
            tk.Label(self.tabla_frame, text=col, bg="#1F4E79", fg="white",
                     font=("Arial", 10, "bold"),
                     relief="flat", pady=8).grid(row=0, column=j,
                                                  sticky="ew", padx=1, pady=(0, 1))

        if not facturas:
            tk.Label(self.tabla_frame,
                     text="No hay facturas registradas aún.",
                     bg="#F0F4F8", fg="#888", font=("Arial", 10),
                     pady=20).grid(row=1, column=0, columnspan=7, sticky="ew")
            return

        for i, f in enumerate(facturas):
            bg = "#DEEAF1" if i % 2 == 0 else "white"
            valores = [f["factura_id"], f["cliente"], f["largo"],
                       f["ancho"], f["area"], f["tarifa"], f"${f['costo']:,.2f}"]
            for j, val in enumerate(valores):
                tk.Label(self.tabla_frame, text=val, bg=bg,
                         font=("Arial", 10),
                         anchor="w" if j == 1 else "center",
                         relief="flat", pady=6).grid(row=i+1, column=j,
                                                      sticky="ew", padx=1, pady=1)

    def _build_footer(self):
        frame = ctk.CTkFrame(self, fg_color="#EEF4FB", corner_radius=0, height=54)
        frame.grid(row=3, column=0, sticky="ew")
        frame.grid_propagate(False)
        frame.grid_columnconfigure(0, weight=1)

        self.lbl_total = ctk.CTkLabel(frame, text="Total recaudado: $0.00",
                                       font=ctk.CTkFont(size=13, weight="bold"),
                                       text_color="#1F4E79")
        self.lbl_total.pack(side="left", padx=20)

        ctk.CTkButton(frame, text="📊 Ver en consola (pandas)",
                      width=200, fg_color="#2E75B6", hover_color="#1F4E79",
                      command=self._mostrar_pandas).pack(side="right", padx=8, pady=10)

        ctk.CTkButton(frame, text="📥 Exportar a Excel",
                      width=160, fg_color="#1F7A3C", hover_color="#155A2D",
                      command=self._exportar_excel).pack(side="right", padx=(0, 4), pady=10)

    def _borrar(self):
            global contador_factura
            if not facturas:
                messagebox.showinfo("Sin datos", "No hay facturas para borrar.")
                return

            opciones = [f"{f['factura_id']} — {f['cliente']}" for f in facturas]
            ventana = ctk.CTkToplevel(self)
            ventana.title("Borrar Factura")
            ventana.geometry("360x220")
            ventana.resizable(False, False)
            ventana.grab_set()

            ctk.CTkLabel(ventana, text="Selecciona la factura a borrar:",
                        font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(20, 8))

            seleccion = ctk.CTkOptionMenu(ventana, values=opciones, width=300)
            seleccion.pack(pady=4)

            def confirmar():
                idx = opciones.index(seleccion.get())
                fid = facturas[idx]["factura_id"]
                if messagebox.askyesno("Confirmar", f"¿Borrar la factura {fid}?"):
                    facturas.pop(idx)
                    self._render_tabla()
                    total = sum(f["costo"] for f in facturas)
                    self.lbl_total.configure(text=f"Total recaudado: ${total:,.2f}")
                    ventana.destroy()

            ctk.CTkButton(ventana, text="Borrar", fg_color="#C0392B",
                        hover_color="#922B21", command=confirmar).pack(pady=16)


    def _registrar(self):
        global contador_factura
        cliente = self.entries["cliente"].get().strip()
        if not cliente:
            messagebox.showwarning("Campo vacío", "Ingresa el nombre del cliente.")
            return
        try:
            largo = float(self.entries["largo"].get())
            ancho = float(self.entries["ancho"].get())
        except ValueError:
            messagebox.showerror("Error", "Largo y ancho deben ser números.")
            return

        area  = largo * ancho
        costo = area * 0.5
        facturas.append({
            "factura_id": f"FAC-{contador_factura:04d}",
            "cliente": cliente, "largo": largo, "ancho": ancho,
            "area": area, "tarifa": 0.5, "costo": costo
        })
        contador_factura += 1

        for e in self.entries.values():
            e.delete(0, "end")

        self._render_tabla()
        total = sum(f["costo"] for f in facturas)
        self.lbl_total.configure(text=f"Total recaudado: ${total:,.2f}")
        messagebox.showinfo("✓ Registrado",
                            f"Factura FAC-{contador_factura-1:04d} — Total: ${costo:,.2f}")

    def _mostrar_pandas(self):
        if not facturas:
            messagebox.showinfo("Sin datos", "No hay facturas registradas.")
            return
        df = pd.DataFrame(facturas)
        df.columns = ["N° Factura", "Cliente", "Largo", "Ancho", "Área", "Tarifa", "Total ($)"]
        df.index += 1
        print("\n" + "="*60)
        print(df.to_string())
        print(f"\nTotal recaudado: ${df['Total ($)'].sum():,.2f}")
        print("="*60)
        messagebox.showinfo("Pandas", "Tabla impresa en la consola ✓")

    def _exportar_excel(self):
        if not facturas:
            messagebox.showinfo("Sin datos", "No hay facturas para exportar.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas Acueducto"

        color_header    = "1F4E79"
        color_subheader = "2E75B6"
        color_fila_par  = "DEEAF1"
        color_total_bg  = "BDD7EE"

        fuente_titulo    = Font(name="Arial", size=16, bold=True,  color="FFFFFF")
        fuente_subtitulo = Font(name="Arial", size=10, italic=True, color="FFFFFF")
        fuente_header    = Font(name="Arial", size=11, bold=True,  color="FFFFFF")
        fuente_normal    = Font(name="Arial", size=10)
        fuente_total     = Font(name="Arial", size=11, bold=True,  color="1F4E79")
        borde = Border(
            left=Side(style="thin",   color="B8CCE4"),
            right=Side(style="thin",  color="B8CCE4"),
            top=Side(style="thin",    color="B8CCE4"),
            bottom=Side(style="thin", color="B8CCE4"))

        ws.merge_cells("A1:G1")
        ws["A1"] = "EMPRESA DE ACUEDUCTO"
        ws["A1"].font      = fuente_titulo
        ws["A1"].fill      = PatternFill("solid", fgColor=color_header)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Reporte de Facturas — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font      = fuente_subtitulo
        ws["A2"].fill      = PatternFill("solid", fgColor=color_subheader)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 8

        encabezados = ["N° Factura", "Cliente", "Largo (m)", "Ancho (m)",
                       "Área (m²)", "Tarifa ($/m²)", "Total ($)"]
        anchos = [13, 28, 13, 13, 13, 15, 14]
        for col, (enc, ancho) in enumerate(zip(encabezados, anchos), 1):
            c = ws.cell(row=4, column=col, value=enc)
            c.font      = fuente_header
            c.fill      = PatternFill("solid", fgColor=color_header)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = borde
            ws.column_dimensions[get_column_letter(col)].width = ancho
        ws.row_dimensions[4].height = 22

        for i, f in enumerate(facturas):
            fila = 5 + i
            bg   = color_fila_par if i % 2 == 0 else "FFFFFF"
            for col, val in enumerate([f["factura_id"], f["cliente"], f["largo"],
                                        f["ancho"], f["area"], f["tarifa"], f["costo"]], 1):
                c = ws.cell(row=fila, column=col, value=val)
                c.font      = fuente_normal
                c.fill      = PatternFill("solid", fgColor=bg)
                c.border    = borde
                c.alignment = Alignment(
                    horizontal="left" if col == 2 else "center", vertical="center")
                if col in (3, 4, 5, 6, 7):
                    c.number_format = "#,##0.00"

        fila_total = 5 + len(facturas)
        ws.merge_cells(f"A{fila_total}:F{fila_total}")
        c = ws[f"A{fila_total}"]
        c.value     = "TOTAL RECAUDADO"
        c.font      = fuente_total
        c.fill      = PatternFill("solid", fgColor=color_total_bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = borde

        ct = ws[f"G{fila_total}"]
        ct.value         = f"=SUM(G5:G{fila_total-1})"
        ct.font          = fuente_total
        ct.fill          = PatternFill("solid", fgColor=color_total_bg)
        ct.alignment     = Alignment(horizontal="center", vertical="center")
        ct.number_format = "#,##0.00"
        ct.border        = borde

        carpeta = os.path.join(os.path.expanduser("~"), "Downloads")
        os.makedirs(carpeta, exist_ok=True)
        nombre = f"Facturas_Acueducto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta   = os.path.join(carpeta, nombre)
        wb.save(ruta)
        messagebox.showinfo("✓ Excel exportado", f"Guardado en:\n{ruta}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
